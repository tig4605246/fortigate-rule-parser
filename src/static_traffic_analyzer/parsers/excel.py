"""Parser for Excel-based firewall rules."""
from __future__ import annotations

from dataclasses import dataclass
from typing import Iterable
import socket

from openpyxl import load_workbook

from ..catalog import DEFAULT_SERVICES
from ..models import AddressBook, AddressGroup, PolicyRule, ServiceBook, ServiceGroup, ServiceObject
from ..utils import ParseError, make_any_service, parse_address_object, parse_service_entry


@dataclass
class ExcelData:
    """Parsed Excel data container."""

    address_book: AddressBook
    service_book: ServiceBook
    policies: list[PolicyRule]


def _split_members(raw_value: str | None) -> list[str]:
    """Split member lists on newlines and commas."""
    if not raw_value:
        return []
    members = []
    for line in str(raw_value).splitlines():
        for part in line.split(","):
            member = part.strip()
            if member:
                members.append(member)
    return members


def parse_excel(path: str) -> ExcelData:
    """Parse Excel workbook according to the specification."""
    workbook = load_workbook(path, data_only=True)

    address_book = AddressBook()
    service_book = ServiceBook()
    policies: list[PolicyRule] = []

    if "Address Object" not in workbook.sheetnames:
        raise ParseError("Missing 'Address Object' sheet in Excel file")
    if "Address Group" not in workbook.sheetnames:
        raise ParseError("Missing 'Address Group' sheet in Excel file")
    if "Service Group" not in workbook.sheetnames:
        raise ParseError("Missing 'Service Group' sheet in Excel file")
    if "Rule" not in workbook.sheetnames:
        raise ParseError("Missing 'Rule' sheet in Excel file")

    address_sheet = workbook["Address Object"]
    headers = [cell.value for cell in next(address_sheet.iter_rows(min_row=1, max_row=1))]
    header_map = {str(header).strip(): idx for idx, header in enumerate(headers) if header is not None}

    for row in address_sheet.iter_rows(min_row=2):
        name = row[header_map.get("Object Name", -1)].value if header_map.get("Object Name") is not None else None
        if not name:
            continue
        address_type = row[header_map.get("Type", -1)].value or "ipmask"
        subnet_value = row[header_map.get("Subnet/Start-IP", -1)].value
        mask_value = row[header_map.get("Mask/End-IP", -1)].value
        subnet = None
        start_ip = None
        end_ip = None
        if str(address_type).lower() == "ipmask" and subnet_value and mask_value:
            subnet = f"{subnet_value}/{mask_value}"
        elif str(address_type).lower() == "iprange":
            start_ip = str(subnet_value) if subnet_value else None
            end_ip = str(mask_value) if mask_value else None
        try:
            address_book.objects[str(name)] = parse_address_object(
                name=str(name),
                address_type=str(address_type),
                subnet=subnet,
                start_ip=start_ip,
                end_ip=end_ip,
            )
        except ParseError:
            address_book.objects[str(name)] = parse_address_object(
                name=str(name),
                address_type="fqdn",
            )

    address_group_sheet = workbook["Address Group"]
    headers = [cell.value for cell in next(address_group_sheet.iter_rows(min_row=1, max_row=1))]
    header_map = {str(header).strip(): idx for idx, header in enumerate(headers) if header is not None}

    for row in address_group_sheet.iter_rows(min_row=2):
        group_name = row[header_map.get("Group Name", -1)].value if header_map.get("Group Name") is not None else None
        member_value = row[header_map.get("Member", -1)].value if header_map.get("Member") is not None else None
        if not group_name:
            continue
        members = tuple(_split_members(member_value))
        address_book.groups[str(group_name)] = AddressGroup(name=str(group_name), members=members)

    service_group_sheet = workbook["Service Group"]
    headers = [cell.value for cell in next(service_group_sheet.iter_rows(min_row=1, max_row=1))]
    header_map = {str(header).strip(): idx for idx, header in enumerate(headers) if header is not None}

    for row in service_group_sheet.iter_rows(min_row=2):
        group_name = row[header_map.get("Group Name", -1)].value if header_map.get("Group Name") is not None else None
        member_value = row[header_map.get("Member", -1)].value if header_map.get("Member") is not None else None
        if not group_name:
            continue
        members = tuple(_split_members(member_value))
        service_book.groups[str(group_name)] = ServiceGroup(name=str(group_name), members=members)

    rule_sheet = workbook["Rule"]
    headers = [cell.value for cell in next(rule_sheet.iter_rows(min_row=1, max_row=1))]
    header_map = {str(header).strip(): idx for idx, header in enumerate(headers) if header is not None}

    for row in rule_sheet.iter_rows(min_row=2):
        seq = row[header_map.get("Seq", -1)].value if header_map.get("Seq") is not None else None
        enable = row[header_map.get("Enable", -1)].value if header_map.get("Enable") is not None else None
        source = row[header_map.get("Source", -1)].value if header_map.get("Source") is not None else None
        destination = row[header_map.get("Destination", -1)].value if header_map.get("Destination") is not None else None
        service = row[header_map.get("Service", -1)].value if header_map.get("Service") is not None else None
        action = row[header_map.get("Action", -1)].value if header_map.get("Action") is not None else None
        rule_id = row[header_map.get("ID", -1)].value if header_map.get("ID") is not None else None
        comments = row[header_map.get("Comments", -1)].value if header_map.get("Comments") is not None else None

        if seq is None:
            continue

        policies.append(
            PolicyRule(
                policy_id=str(rule_id) if rule_id is not None else str(seq),
                name=str(rule_id) if rule_id is not None else str(seq),
                priority=int(seq),
                source=tuple(_split_members(source)),
                destination=tuple(_split_members(destination)),
                services=tuple(_split_members(service)),
                action=str(action or "deny"),
                enabled=str(enable).lower() == "true",
                schedule="always",
                comment=str(comments) if comments else None,
            )
        )

    for name, service in DEFAULT_SERVICES.items():
        service_book.services.setdefault(name, service)
    if "ALL" not in service_book.services:
        service_book.services["ALL"] = make_any_service("ALL")

    for group in list(service_book.groups.values()):
        for member in group.members:
            if member in service_book.services:
                continue
            if member.lower().startswith("tcp_") or member.lower().startswith("udp_"):
                try:
                    service_book.services[member] = ServiceObject(name=member, entries=(parse_service_entry(member),))
                except ParseError:
                    continue
            else:
                try:
                    service_book.services[member] = ServiceObject(name=member, entries=(socket.getservbyname(member),))
                except (OSError, TypeError):
                    continue

    policies.sort(key=lambda rule: rule.priority)

    return ExcelData(address_book=address_book, service_book=service_book, policies=policies)

"""Tests for Excel parser edge cases."""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from static_traffic_analyzer.models import AddressObject
from static_traffic_analyzer.parsers.excel import ParseError, parse_excel
from static_traffic_analyzer.utils import parse_address_object


@pytest.fixture
def base_workbook(tmp_path: Path) -> Path:
    """Creates a valid base workbook structure for testing."""
    workbook = Workbook()
    workbook.remove(workbook.active)  # Remove default sheet

    # Create all required sheets
    address_sheet = workbook.create_sheet("Address Object")
    address_sheet.append(["Object Name", "Type", "Subnet/Start-IP", "Mask/End-IP"])
    address_sheet.append(["net", "ipmask", "10.0.0.0", "255.255.255.0"])
    address_sheet.append(["host", "fqdn", "example.com", ""])

    address_group_sheet = workbook.create_sheet("Address Group")
    address_group_sheet.append(["Group Name", "Member"])
    address_group_sheet.append(["group", "net,host"])

    service_group_sheet = workbook.create_sheet("Service Group")
    service_group_sheet.append(["Group Name", "Member"])
    service_group_sheet.append(["svc-group", "tcp_80, udp_53"])

    rule_sheet = workbook.create_sheet("Rule")
    rule_sheet.append(["Seq", "Enable", "Source", "Destination", "Service", "Action", "ID", "Comments"])
    rule_sheet.append([2, "FALSE", "group", "any", "svc-group", "deny", "RULE-B", "Comment B"])
    rule_sheet.append([1, "TRUE", "any", "group", "tcp_443", "accept", "RULE-A", "Comment A"])

    path = tmp_path / "rules.xlsx"
    workbook.save(path)
    return path


def test_parse_excel_success(base_workbook: Path):
    """Test successful parsing of a valid Excel file."""
    data = parse_excel(str(base_workbook))

    # Address Objects
    assert data.address_book.objects["net"] == parse_address_object(name="net", address_type="ipmask", subnet="10.0.0.0/255.255.255.0")
    assert data.address_book.objects["host"] == parse_address_object(name="host", address_type="fqdn")

    # Address Groups
    assert data.address_book.groups["group"].members == ("net", "host")

    # Service Groups
    assert data.service_book.groups["svc-group"].members == ("tcp_80", "udp_53")

    # Policies (should be sorted by Seq)
    assert len(data.policies) == 2
    policy1, policy2 = data.policies

    assert policy1.priority == 1
    assert policy1.policy_id == "RULE-A"
    assert policy1.enabled is True
    assert policy1.source == ("any",)
    assert policy1.destination == ("group",)
    assert policy1.services == ("tcp_443",)
    assert policy1.action == "accept"

    assert policy2.priority == 2
    assert policy2.policy_id == "RULE-B"
    assert policy2.enabled is False
    assert policy2.source == ("group",)
    assert policy2.destination == ("any",)
    assert policy2.services == ("svc-group",)
    assert policy2.action == "deny"


@pytest.mark.parametrize(
    "missing_sheet",
    ["Address Object", "Address Group", "Service Group", "Rule"],
)
def test_excel_missing_sheet(tmp_path: Path, missing_sheet: str):
    """Test that a ParseError is raised if a required sheet is missing."""
    workbook = Workbook()
    # Create all sheets except the one to be tested
    for sheet_name in ["Address Object", "Address Group", "Service Group", "Rule"]:
        if sheet_name != missing_sheet:
            workbook.create_sheet(sheet_name)

    path = tmp_path / "rules.xlsx"
    workbook.save(path)

    with pytest.raises(ParseError, match=f"Missing '{missing_sheet}' sheet"):
        parse_excel(str(path))


def test_excel_member_splitting(tmp_path: Path):
    """Test various member splitting scenarios."""
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.create_sheet("Address Object")  # Needs to exist
    workbook.create_sheet("Service Group")  # Needs to exist
    workbook.create_sheet("Rule")  # Needs to exist

    address_group_sheet = workbook.create_sheet("Address Group")
    address_group_sheet.append(["Group Name", "Member"])
    address_group_sheet.append(["group1", " member1, member2 "])
    address_group_sheet.append(["group2", "member3\nmember4"])
    address_group_sheet.append(["group3", "member5,member6\nmember7"])
    address_group_sheet.append(["group4", " "])  # Empty/whitespace only
    address_group_sheet.append(["group5", None])  # None value

    path = tmp_path / "rules.xlsx"
    workbook.save(path)

    data = parse_excel(str(path))
    assert data.address_book.groups["group1"].members == ("member1", "member2")
    assert data.address_book.groups["group2"].members == ("member3", "member4")
    assert data.address_book.groups["group3"].members == ("member5", "member6", "member7")
    assert data.address_book.groups["group4"].members == ()
    assert data.address_book.groups["group5"].members == ()


def test_excel_missing_headers(tmp_path: Path):
    """Test that parsing continues gracefully when some headers are missing."""
    workbook = Workbook()
    workbook.remove(workbook.active)

    # Missing "Type" header
    address_sheet = workbook.create_sheet("Address Object")
    address_sheet.append(["Object Name", "Subnet/Start-IP", "Mask/End-IP"])
    address_sheet.append(["net", "10.0.0.0", "255.255.255.0"])

    # Missing "Member" header
    address_group_sheet = workbook.create_sheet("Address Group")
    address_group_sheet.append(["Group Name"])
    address_group_sheet.append(["group"])

    workbook.create_sheet("Service Group")
    workbook.create_sheet("Rule")

    path = tmp_path / "rules.xlsx"
    workbook.save(path)

    data = parse_excel(str(path))
    # Falls back to ipmask by default, but can't form subnet without mask/end-ip from header.
    # It will try to parse and then fall back to fqdn.
    assert data.address_book.objects["net"] == parse_address_object(name="net", address_type="fqdn")
    assert data.address_book.groups["group"].members == ()


def test_excel_member_empty_lines(tmp_path: Path):
    workbook = Workbook()
    workbook.remove(workbook.active)

    address_sheet = workbook.create_sheet("Address Object")
    address_sheet.append(["Object Name", "Type", "Subnet/Start-IP", "Mask/End-IP"])
    address_sheet.append(["net", "ipmask", "10.0.0.0", "255.255.255.0"])

    address_group_sheet = workbook.create_sheet("Address Group")
    address_group_sheet.append(["Group Name", "Member"])
    address_group_sheet.append(["group", "net\n\n"])

    service_group_sheet = workbook.create_sheet("Service Group")
    service_group_sheet.append(["Group Name", "Member"])
    service_group_sheet.append(["svc-group", "tcp_80\n\n"])

    rule_sheet = workbook.create_sheet("Rule")
    rule_sheet.append(["Seq", "Enable", "Source", "Destination", "Service", "Action", "ID", "Comments"])
    rule_sheet.append([1, True, "group", "group", "svc-group", "accept", "1", "test"])

    path = tmp_path / "rules.xlsx"
    workbook.save(path)

    data = parse_excel(str(path))

    assert data.address_book.groups["group"].members == ("net",)
    assert data.service_book.groups["svc-group"].members == ("tcp_80",)
    assert data.policies[0].services == ("svc-group",)
